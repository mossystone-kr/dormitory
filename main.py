import sys
import os
import PyQt5
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5 import uic
import pandas as pd
import random as rd
import copy
import webbrowser
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill

def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
form_class = uic.loadUiType(BASE_DIR + r'\design.ui')[0]

# 사람, 방 리스트 생성
freshman_list = []
junior_list = []
senior_list = []
freshman_m_list = []
junior_m_list = []
senior_m_list = []
freshman_f_list = []
junior_f_list = []
senior_f_list = []
freshman_m_room = []
junior_m_room = []
senior_m_room = []
freshman_f_room = []
junior_f_room = []
senior_f_room = []
seat_list = []
list_to_fix_room = []
list_to_fix_seat = []
freshman_num = junior_num = senior_num = 0
student_xlsx = 0

# 값 참조만 하면 상관없는데 바꿔야 할 경우를 대비해서 아래에 복붙할 글로벌 키워드를 적어놓음
# global freshman_list, freshman_m_list, freshman_f_list, freshman_m_room, freshman_f_room, freshman_num
# global junior_list, junior_m_list, junior_f_list, junior_m_room, junior_f_room, junior_num
# global senior_list, senior_m_list, senior_f_list, senior_f_room, senior_m_room, senior_num
# global seat_list, list_to_fix_room, list_to_fix_seat

# 방 리스트를 pop해서 배치하니까 복구를 해줘야 함(나중에 최적화)
def getRoom():
    global freshman_m_room, freshman_f_room, junior_m_room, junior_f_room, senior_m_room, senior_f_room
    freshman_m_room=[]
    freshman_f_room=[]
    junior_m_room=[]
    junior_f_room=[]
    senior_m_room=[]
    senior_f_room=[]
    for i in range(31):
        if int(student_xlsx["학년1"][i]) == 1:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            freshman_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 2:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            junior_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 3:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            senior_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 4:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            freshman_f_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 5:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            junior_f_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 6:
            b = [201 + i, int(student_xlsx["인원수1"][i])]
            senior_f_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 7:
            b = [201 + i, 1]
            freshman_m_room.append(b)
            b = [201 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 8:
            b = [201 + i, 1]
            senior_m_room.append(b)
            b = [201 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 9:
            b = [201 + i, 1]
            freshman_m_room.append(b)
            b = [201 + i, 1]
            senior_m_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 10:
            b = [201 + i, 1]
            freshman_f_room.append(b)
            b = [201 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 11:
            b = [201 + i, 1]
            senior_f_room.append(b)
            b = [201 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년1"][i]) == 12:
            b = [201 + i, 1]
            freshman_f_room.append(b)
            b = [201 + i, 1]
            senior_f_room.append(b)
    for i in range(31):
        if int(student_xlsx["학년2"][i]) == 1:
            b = [201 + i, int(student_xlsx["인원수2"][i])]
            freshman_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 2:
            b = [301 + i, int(student_xlsx["인원수2"][i])]
            junior_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 3:
            b = [301 + i, int(student_xlsx["인원수2"][i])]
            senior_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 4:
            b = [301 + i, int(student_xlsx["인원수2"][i])]
            freshman_f_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 5:
            b = [301 + i, int(student_xlsx["인원수2"][i])]
            junior_f_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 6:
            b = [301 + i, int(student_xlsx["인원수2"][i])]
            senior_f_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 7:
            b = [301 + i, 1]
            freshman_m_room.append(b)
            b = [301 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 8:
            b = [301 + i, 1]
            senior_m_room.append(b)
            b = [301 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 9:
            b = [301 + i, 1]
            freshman_m_room.append(b)
            b = [301 + i, 1]
            senior_m_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 10:
            b = [301 + i, 1]
            freshman_f_room.append(b)
            b = [301 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 11:
            b = [301 + i, 1]
            senior_f_room.append(b)
            b = [301 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년2"][i]) == 12:
            b = [301 + i, 1]
            freshman_f_room.append(b)
            b = [301 + i, 1]
            senior_f_room.append(b)
    for i in range(26):
        if int(student_xlsx["학년3"][i]) == 1:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            freshman_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 2:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            junior_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 3:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            senior_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 4:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            freshman_f_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 5:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            junior_f_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 6:
            b = [401 + i, int(student_xlsx["인원수3"][i])]
            senior_f_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 7:
            b = [401 + i, 1]
            freshman_m_room.append(b)
            b = [401 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 8:
            b = [401 + i, 1]
            senior_m_room.append(b)
            b = [401 + i, 1]
            junior_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 9:
            b = [401 + i, 1]
            freshman_m_room.append(b)
            b = [401 + i, 1]
            senior_m_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 10:
            b = [401 + i, 1]
            freshman_f_room.append(b)
            b = [401 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 11:
            b = [401 + i, 1]
            senior_f_room.append(b)
            b = [401 + i, 1]
            junior_f_room.append(b)
        elif int(student_xlsx["학년3"][i]) == 12:
            b = [401 + i, 1]
            freshman_f_room.append(b)
            b = [401 + i, 1]
            senior_f_room.append(b)

# 학생 클래스
class Student:
    def __init__(self):
        self.num = 0
        self.name = "g"
        self.sex = 0
        self.room = 0
        self.grade = 0
        self.seat = [0, 0]

# 화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.save = [False] * 7
        self.btn_1g.clicked.connect(lambda: self.roomPlan1(self.save[2]))
        self.btn_2g.clicked.connect(lambda: self.roomPlan2(self.save[3]))
        self.btn_3g.clicked.connect(lambda: self.roomPlan3(self.save[4]))
        self.btn_sm.clicked.connect(lambda: self.studyPlanM(self.save[5]))
        self.btn_sf.clicked.connect(lambda: self.studyPlanF(self.save[6]))

        self.check_1g.stateChanged.connect(lambda: self.checkFunction(2))
        self.check_2g.stateChanged.connect(lambda: self.checkFunction(3))
        self.check_3g.stateChanged.connect(lambda: self.checkFunction(4))
        self.check_sm.stateChanged.connect(lambda: self.checkFunction(5))
        self.check_sf.stateChanged.connect(lambda: self.checkFunction(6))

        self.map_1.clicked.connect(lambda: self.showMap(1))
        self.map_2.clicked.connect(lambda: self.showMap(2))
        self.map_3.clicked.connect(lambda: self.showMap(3))

        self.saveMode = 0  # 0: 층별 / 1: 학년별
        self.rBtn_floor.toggled.connect(lambda: self.changeMode(0))
        self.rBtn_grade.toggled.connect(lambda: self.changeMode(1))
        self.btn_xlsx.clicked.connect(lambda: self.exportXLSX(self.saveMode))
        self.btn_txt.clicked.connect(lambda: self.exportTXT(self.saveMode))
        self.actionExit.triggered.connect(qApp.quit)
        self.studentFileName = ''
        self.actionUpload.triggered.connect(self.openFile)
        self.plan = [[] for i in range(5)]
        self.tmpPlan = [[] for i in range(5)]
        self.actionHelp.triggered.connect(lambda: webbrowser.open('https://docs.google.com/document/d/1vIYdPUKljWS9jPG89pcUGDy9y6tLLtvv6ehw801yguo/edit?usp=sharing'))

    def changeMode(self, n):
        self.saveMode = n

    def roomPlan1(self, ifsave):
        global freshman_list, freshman_m_list, freshman_f_list, freshman_m_room, freshman_f_room, freshman_num
        global seat_list, list_to_fix_room, list_to_fix_seat
        if self.studentFileName == '':
            msg=QMessageBox()
            msg.setText('No File Selected')
            msg.setWindowTitle('Warning')
            msg.exec_()
            return
        if ifsave:
            self.plan[0] = copy.deepcopy(self.tmpPlan[0])
        self.tmpPlan[0].clear()
        getRoom()
        rd.shuffle(freshman_list)
        for j in range(freshman_num):
            student = freshman_list[j]
            if student.sex == 0:
                student.room = freshman_m_room[0][0]
                freshman_m_room[0][1] -= 1
                if freshman_m_room[0][1] == 0: freshman_m_room.pop(0)
            else:
                student.room = freshman_f_room[0][0]
                freshman_f_room[0][1] -= 1
                if freshman_f_room[0][1] == 0: freshman_f_room.pop(0)
        x=sorted(freshman_list, key=lambda x: x.room)
        for tmp in x:
            self.tmpPlan[0].append(tmp)
        str_x=''
        self.lbl_1g.clear()
        for tmpStudent in x:
            tmp=(tmpStudent.name+' '+str(tmpStudent.room)+'\n')
            str_x+=tmp
            self.lbl_1g.append(tmp)

    def roomPlan2(self, ifsave):
        global junior_list, junior_m_list, junior_f_list, junior_m_room, junior_f_room, junior_num
        global seat_list, list_to_fix_room, list_to_fix_seat
        if self.studentFileName == '':
            QMessageBox.about(self, 'Warning', 'No File Selected')
            return
        if ifsave:
            self.plan[1] = copy.deepcopy(self.tmpPlan[1])
        self.tmpPlan[1].clear()
        getRoom()
        rd.shuffle(junior_list)
        for j in range(junior_num):
            student = junior_list[j]
            if student.sex == 0:
                student.room = junior_m_room[0][0]
                junior_m_room[0][1] -= 1
                if junior_m_room[0][1] == 0: junior_m_room.pop(0)
            else:
                student.room = junior_f_room[0][0]
                junior_f_room[0][1] -= 1
                if junior_f_room[0][1] == 0: junior_f_room.pop(0)
        x = sorted(junior_list, key=lambda x: x.room)
        for tmp in x:
            self.tmpPlan[1].append(tmp)
        str_x = ''
        self.lbl_2g.setText('')
        for tmpStudent in x:
            tmp = (tmpStudent.name + ' ' + str(tmpStudent.room) + '\n')
            str_x += tmp
            self.lbl_2g.append(tmp)

    def roomPlan3(self, ifsave):
        global senior_list, senior_m_list, senior_f_list, senior_f_room, senior_m_room, senior_num
        global seat_list, list_to_fix_room, list_to_fix_seat
        if self.studentFileName == '':
            QMessageBox.about(self, 'Warning', 'No File Selected')
            return
        if ifsave:
            self.plan[2] = copy.deepcopy(self.tmpPlan[2])
        self.tmpPlan[2].clear()
        getRoom()
        rd.shuffle(senior_list)
        for j in range(senior_num):
            student = senior_list[j]
            if student.sex == 0:
                student.room = senior_m_room[0][0]
                senior_m_room[0][1] -= 1
                if senior_m_room[0][1] == 0: senior_m_room.pop(0)
            else:
                student.room = senior_f_room[0][0]
                senior_f_room[0][1] -= 1
                if senior_f_room[0][1] == 0: senior_f_room.pop(0)
        x = sorted(senior_list, key=lambda x: x.room)
        for tmp in x:
            self.tmpPlan[2].append(tmp)
        str_x = ''
        self.lbl_3g.setText('')
        for tmpStudent in x:
            tmp = (tmpStudent.name + ' ' + str(tmpStudent.room) + '\n')
            str_x += tmp
            self.lbl_3g.append(tmp)

    def studyPlanM(self, ifsave):
        global freshman_m_list, junior_m_list, senior_m_list
        global seat_list, list_to_fix_seat
        if self.studentFileName == '':
            QMessageBox.about(self, 'Warning', 'No File Selected')
            return
        if ifsave:
            self.plan[3] = copy.deepcopy(self.tmpPlan[3])
        self.tmpPlan[3].clear()
        count=[0,0,0]
        rd.shuffle(freshman_m_list)
        rd.shuffle(junior_m_list)
        rd.shuffle(senior_m_list)
        for i in range(15):
            k = "북쪽라인" + str(i + 1)
            a = student_xlsx[k]
            lis = []
            for m in range(15):
                if a[2 + m] == 0:
                    lis.append(0)
                elif a[2 + m] == 1:
                    freshman_m_list[count[0]].seat=[i,m]
                    count[0]+=1
                elif a[2 + m] == 2:
                    junior_m_list[count[1]].seat = [i, m]
                    count[1] += 1
                elif a[2 + m] == 3:
                    senior_m_list[count[2]].seat = [i, m]
                    count[2] += 1
            seat_list.append(lis)
        x = sorted(freshman_m_list, key=lambda x: x.seat)
        y = sorted(junior_m_list, key=lambda x: x.seat)
        z = sorted(senior_m_list, key=lambda x: x.seat)
        for tmp in x:
            self.tmpPlan[3].append(tmp)
        for tmp in y:
            self.tmpPlan[3].append(tmp)
        for tmp in z:
            self.tmpPlan[3].append(tmp)
        li=x+y+z
        str_li=''
        self.lbl_sm.setText('')
        for tmpStudent in li:
            tmp = (tmpStudent.name + ' ' + str(tmpStudent.seat) + '\n')
            str_li += tmp
            self.lbl_sm.append(tmp)

    def studyPlanF(self, ifsave):
        global freshman_f_list, junior_f_list, senior_f_list
        global seat_list, list_to_fix_seat
        if self.studentFileName == '':
            QMessageBox.about(self, 'Warning', 'No File Selected')
            return
        if ifsave:
            self.plan[4] = copy.deepcopy(self.tmpPlan[4])
        self.tmpPlan[4].clear()
        count = [0, 0, 0]
        seat_list=[]
        rd.shuffle(freshman_f_list)
        rd.shuffle(junior_f_list)
        rd.shuffle(senior_f_list)
        for i in range(15):
            k = "북쪽라인" + str(i + 1)
            a = student_xlsx[k]
            lis = []
            for m in range(15):
                if a[2 + m] == 0:
                    lis.append(0)
                elif a[2 + m] == 4:
                    freshman_f_list[count[0]].seat = [i, m]
                    count[0] += 1
                elif a[2 + m] == 5:
                    junior_f_list[count[1]].seat = [i, m]
                    count[1] += 1
                elif a[2 + m] == 6:
                    senior_f_list[count[2]].seat = [i, m]
                    count[2] += 1
            seat_list.append(lis)
        x = sorted(freshman_f_list, key=lambda x: x.seat)
        y = sorted(junior_f_list, key=lambda x: x.seat)
        z = sorted(senior_f_list, key=lambda x: x.seat)
        for tmp in x:
            self.tmpPlan[4].append(tmp)
        for tmp in y:
            self.tmpPlan[4].append(tmp)
        for tmp in z:
            self.tmpPlan[4].append(tmp)
        li = x + y + z
        str_li = ''
        self.lbl_sf.setText('')
        for tmpStudent in li:
            tmp = (tmpStudent.name + ' ' + str(tmpStudent.seat) + '\n')
            str_li += tmp
            self.lbl_sf.append(tmp)


    def exportXLSX(self, saveMode):
        wb = openpyxl.Workbook()
        alpha_list = [chr(i) for i in range(65, 80)]
        grade_color = [0, 'FFFF5A', '77FF70', '6799FF']
        tSide = Side(border_style='thin')
        # 가로: A, B, C, D ...
        # 세로: 1, 2, 3, 4 ...
        ws_seat = wb.create_sheet('자습실')
        del wb['Sheet']
        ws_room = [wb.create_sheet('1학년'), wb.create_sheet('2학년'), wb.create_sheet('3학년')]

        for i in range(3):
            ws_room[i]['A1'] = '방'
            ws_room[i]['B1'] = '이름'
            ws_room[i]['A1'].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
            ws_room[i]['B1'].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)

            ws_seat[alpha_list[i] + '18'] = str(i + 1) + '학년'
            ws_seat[alpha_list[i] + '18'].fill = PatternFill(start_color=grade_color[i + 1], fill_type='solid')

        saveplan=[0]*5
        for i in range(5):
            if not self.plan[i]: # plan에 원소 X
                saveplan[i]=copy.deepcopy(self.tmpPlan[i])
            else:
                saveplan[i]=copy.deepcopy(self.plan[i])

        for i in range(3,5):
            for j, tmp in enumerate(saveplan[i]):
                coord = tmp.seat
                xlcoord = alpha_list[coord[0]] + str(coord[1] + 1)
                ws_seat[xlcoord] = tmp.name
                ws_seat[xlcoord].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
                ws_seat[xlcoord].fill = PatternFill(start_color=grade_color[tmp.grade], fill_type='solid')

        if saveMode==1:
            tmpRoom = 0
            for i in range(3):
                for j, tmp in enumerate(saveplan[i]):
                    ws_room[i]['A' + str(j + 2)] = tmp.room
                    ws_room[i]['B' + str(j + 2)] = tmp.name
                    if tmp.room == tmpRoom:
                        ws_room[i].merge_cells('A' + str(j + 1) + ':A' + str(j + 2))
                    tmpRoom = tmp.room
                    ws_room[i]['A' + str(j + 2)].alignment = Alignment(horizontal='right', vertical='center')
                    ws_room[i]['B' + str(j + 2)].alignment = Alignment(horizontal='left')
                    ws_room[i]['A' + str(j + 2)].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
                    ws_room[i]['B' + str(j + 2)].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
        else:
            ws_room[0].title = '2층'
            ws_room[1].title = '3층'
            ws_room[2].title = '4층'
            tmpRoom=0
            tmpIndex=[1,2,3]
            for i in range(3):
                for j, tmp in enumerate(saveplan[i]):
                    if tmp.room >= 400:
                        tmpIndex[i]=j
                        break
                    k=i
                    if i==2:
                        k=1
                        j+=tmpIndex[1]
                    ws_room[k]['A' + str(j + 2)] = tmp.room
                    ws_room[k]['B' + str(j + 2)] = tmp.name
                    if tmp.room == tmpRoom:
                        ws_room[k].merge_cells('A' + str(j + 1) + ':A' + str(j + 2))
                    tmpRoom = tmp.room
                    ws_room[k]['A' + str(j + 2)].alignment = Alignment(horizontal='right', vertical='center')
                    ws_room[k]['B' + str(j + 2)].alignment = Alignment(horizontal='left')
                    ws_room[k]['A' + str(j + 2)].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
                    ws_room[k]['B' + str(j + 2)].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
            savepoint=0
            for i in range(3):
                for j in range(tmpIndex[i], len(saveplan[i])):
                    tmp = saveplan[i][j]
                    ind = str(j+2-tmpIndex[i]+savepoint)
                    ws_room[2]['A' + ind] = tmp.room
                    ws_room[2]['B' + ind] = tmp.name
                    if tmp.room == tmpRoom:
                        ws_room[2].merge_cells('A' + str(int(ind)-1) + ':A' + ind)
                    tmpRoom = tmp.room
                    ws_room[2]['A' + ind].alignment = Alignment(horizontal='right', vertical='center')
                    ws_room[2]['B' + ind].alignment = Alignment(horizontal='left')
                    ws_room[2]['A' + ind].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
                    ws_room[2]['B' + ind].border = Border(top=tSide, right=tSide, bottom=tSide, left=tSide)
                savepoint+=len(saveplan[i])-tmpIndex[i]
        # 저장하기
        wb.save('placement.xlsx')

    def exportTXT(self, saveMode):
        f=open('배치.txt', 'w')
        saveplan=[0]*5
        for i in range(5):
            if not self.plan[i]:
                saveplan[i]=copy.deepcopy(self.tmpPlan[i])
            else:
                saveplan[i]=copy.deepcopy(self.plan[i])

        if saveMode == 1:
            for i in range(5):
                for tmp in saveplan[i]:
                    if i <= 2:
                        f.write(tmp.name + ' ' + str(tmp.room) + '\n')
                    else:
                        f.write(tmp.name + ' ' + str(tmp.seat) + '\n')
                f.write('\n')
        else:
            tmpIndex=[0]*5
            for i in range(3):
                for j, tmp in enumerate(saveplan[i]):
                    if tmp.room >= 400:
                        tmpIndex[i]=j
                        break
                    f.write(tmp.name + ' ' + str(tmp.room) + '\n')
                f.write('\n')
            for i in range(3):
                for j in range(tmpIndex[i], len(saveplan[i])):
                    tmp=saveplan[i][j]
                    f.write(tmp.name + ' ' + str(tmp.room) + '\n')
            f.write('\n')
            for i in range(3,5):
                for tmp in saveplan[i]:
                    f.write(tmp.name + ' ' + str(tmp.seat) + '\n')
                f.write('\n')
        f.close()

    def checkFunction(self, num):  # 2, 3, 4: 층 / 5: 자습실 남 / 6: 자습실 여
        if num == 2:
            if self.check_2f.isChecked():
                self.save[2] = True
            else:
                self.save[2] = False
        elif num == 3:
            if self.check_3f.isChecked():
                self.save[3] = True
            else:
                self.save[3] = False
        elif num == 4:
            if self.check_4f.isChecked():
                self.save[4] = True
            else:
                self.save[4] = False
        elif num == 5:
            if self.check_sm.isChecked():
                self.save[5] = True
            else:
                self.save[5] = False
        elif num == 6:
            if self.check_sf.isChecked():
                self.save[6] = True
            else:
                self.save[6] = False

    def showMap(self, num):
        if num==1:
            self.pic.setPixmap(QPixmap('pic/floor2.png'))
        elif num==2:
            self.pic.setPixmap(QPixmap('pic/floor3.png'))
        elif num==3:
            self.pic.setPixmap(QPixmap('pic/floor4.png'))

    def openFile(self):
        global freshman_num, junior_num, senior_num, student_xlsx
        global freshman_list, freshman_m_list, freshman_f_list, junior_list, junior_m_list, junior_f_list, senior_list, senior_m_list, senior_f_list
        filename = QFileDialog.getOpenFileName(self, 'Open File')
        if filename[0] != '':
            self.studentFileName = filename[0]
        else:
            QMessageBox.about(self, 'Warning', 'No File Selected')
        student_xlsx = pd.read_excel(self.studentFileName)  # 이 부분을 studentFileName을 이용해서 바꿔주면 됨
        freshman_num = int(student_xlsx["1학년 전체 수"][0])
        junior_num = int(student_xlsx["2학년 전체 수"][0])
        senior_num = int(student_xlsx["3학년 전체 수"][0])

        # 학생 정보 추출
        for i in range(freshman_num):
            a = Student()
            a.num = int(student_xlsx["학번1"][i])
            a.name = student_xlsx["이름1"][i]
            a.grade = 1
            if student_xlsx["성별1"][i] == "남":
                a.sex = 0
                freshman_m_list.append(a)
            else:
                a.sex = 1
                freshman_f_list.append(a)
        for i in range(junior_num):
            a = Student()
            a.num = int(student_xlsx["학번2"][i])
            a.name = student_xlsx["이름2"][i]
            a.grade = 2
            if student_xlsx["성별2"][i] == "남":
                a.sex = 0
                junior_m_list.append(a)
            else:
                a.sex = 1
                junior_f_list.append(a)
            junior_list.append(a)
        for i in range(senior_num):
            a = Student()
            a.num = int(student_xlsx["학번3"][i])
            a.name = student_xlsx["이름3"][i]
            a.grade = 3
            if student_xlsx["성별3"][i] == "남":
                a.sex = 0
                senior_m_list.append(a)
            else:
                a.sex = 1
                senior_f_list.append(a)
            senior_list.append(a)

        freshman_list = freshman_m_list + freshman_f_list
        junior_list = junior_m_list + junior_f_list
        senior_list = senior_m_list + senior_f_list

        # 배치할 방 추출
        getRoom()


def exception_hook(exctype, value, traceback):
    print(exctype, value, traceback)
    sys._excepthook(exctype, value, traceback)
    sys.exit(1)


if __name__ == "__main__":
    sys._excepthook = sys.excepthook
    sys.excepthook = exception_hook
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QApplication(sys.argv)
    app.setAttribute(Qt.AA_EnableHighDpiScaling)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
